from flask import Blueprint, render_template, request, send_file, flash, redirect, url_for
from flask_login import login_required
from extensions import mysql
from datetime import date
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

descargas_bp = Blueprint('descargas', __name__, url_prefix='/descargas')

# ── helpers ────────────────────────────────────────────────────────────────────

def get_config():
    cur = mysql.connection.cursor()
    cur.execute("SELECT nombre_persona1, nombre_persona2 FROM configuracion WHERE id = 1")
    cfg = cur.fetchone()
    cur.close()
    return cfg


def meses_disponibles():
    inicio = date(2024, 12, 1)
    hoy    = date.today().replace(day=1)
    meses  = []
    cur    = inicio
    while cur <= hoy:
        meses.append({'year': cur.year, 'month': cur.month,
                      'label': cur.strftime('%B %Y')})
        if cur.month == 12:
            cur = date(cur.year + 1, 1, 1)
        else:
            cur = date(cur.year, cur.month + 1, 1)
    return meses


def datos_mes(year, month):
    mes = date(year, month, 1)
    cur = mysql.connection.cursor()

    # Gastos cuya cuota cae en este mes
    cur.execute("""
        SELECT id, descripcion, categoria, pagado_por, monto_total,
               meses_diferidos, mes_inicio, fecha_registro, notas
        FROM gastos
        WHERE mes_inicio <= %s
          AND DATE_ADD(mes_inicio, INTERVAL (meses_diferidos - 1) MONTH) >= %s
        ORDER BY mes_inicio, id
    """, (mes, mes))
    gastos_raw = cur.fetchall()

    # Pagos extra
    cur.execute("""
        SELECT descripcion, monto, pagado_por, recibido_por, notas, fecha_registro
        FROM pagos_extra
        WHERE mes = %s
        ORDER BY fecha_registro
    """, (mes,))
    pagos_extra = cur.fetchall()
    cur.close()

    gastos = []
    for g in gastos_raw:
        cuota = round(float(g['monto_total']) / g['meses_diferidos'], 2)
        gastos.append({**g, 'cuota_mes': cuota, 'mitad': round(cuota / 2, 2),
                       'monto_total': float(g['monto_total'])})

    return gastos, pagos_extra


# ── Colores ────────────────────────────────────────────────────────────────────
COLOR_HEADER    = '2D6A4F'   # verde oscuro
COLOR_SUBHEADER = '52B788'   # verde medio
COLOR_TOTAL     = 'D8F3DC'   # verde claro
COLOR_WHITE     = 'FFFFFF'
COLOR_DEUDA     = 'FFF3CD'   # amarillo suave
COLOR_ALT       = 'F9F7F4'   # gris muy suave

BORDER_THIN = Border(
    left  =Side(style='thin', color='D0CCBF'),
    right =Side(style='thin', color='D0CCBF'),
    top   =Side(style='thin', color='D0CCBF'),
    bottom=Side(style='thin', color='D0CCBF'),
)

def hdr_font(bold=True, color=COLOR_WHITE, size=10):
    return Font(name='Arial', bold=bold, color=color, size=size)

def body_font(bold=False, color='1A1814', size=10):
    return Font(name='Arial', bold=bold, color=color, size=size)

def fill(hex_color):
    return PatternFill('solid', start_color=hex_color, fgColor=hex_color)

def center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def vcenter():
    return Alignment(horizontal='left', vertical='center', wrap_text=True)

def money_fmt():
    return '$#,##0.00'

def apply_border(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                             min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = BORDER_THIN


# ── Construcción del Excel ─────────────────────────────────────────────────────

def build_excel(year, month, gastos, pagos_extra, cfg):
    n1 = cfg['nombre_persona1']
    n2 = cfg['nombre_persona2']
    mes_label = date(year, month, 1).strftime('%B %Y').capitalize()

    wb = openpyxl.Workbook()

    # ── Hoja 1: Resumen ────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Resumen'
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

    # Título
    ws.merge_cells('A1:D1')
    ws['A1'] = f'Reporte de Gastos — {mes_label}'
    ws['A1'].font = Font(name='Arial', bold=True, color=COLOR_WHITE, size=14)
    ws['A1'].fill = fill(COLOR_HEADER)
    ws['A1'].alignment = center()
    ws.row_dimensions[1].height = 32

    # Subtítulo
    ws.merge_cells('A2:D2')
    ws['A2'] = f'Generado el {date.today().strftime("%d/%m/%Y")}'
    ws['A2'].font = Font(name='Arial', color='6B6860', size=9, italic=True)
    ws['A2'].fill = fill('F7F5F2')
    ws['A2'].alignment = center()
    ws.row_dimensions[2].height = 16

    ws.append([])

    # Encabezados balance
    headers_bal = ['', n1, n2, 'Total']
    ws.append(headers_bal)
    r = ws.max_row
    for c, h in enumerate(headers_bal, 1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = hdr_font()
        cell.fill = fill(COLOR_SUBHEADER)
        cell.alignment = center()
        cell.border = BORDER_THIN
    ws.row_dimensions[r].height = 20

    # Calcula totales
    total_p1 = sum(g['cuota_mes'] for g in gastos if g['pagado_por'] == 'persona1')
    total_p2 = sum(g['cuota_mes'] for g in gastos if g['pagado_por'] == 'persona2')
    extra_p1 = sum(float(p['monto']) for p in pagos_extra if p['pagado_por'] == 'persona1')
    extra_p2 = sum(float(p['monto']) for p in pagos_extra if p['pagado_por'] == 'persona2')
    total_gastado = total_p1 + total_p2
    mitad = round(total_gastado / 2, 2)
    balance_p1 = round(total_p1 - mitad + extra_p1 - extra_p2, 2)
    balance_p2 = round(total_p2 - mitad + extra_p2 - extra_p1, 2)

    filas_bal = [
        ('Gastos pagados',       total_p1, total_p2, total_gastado),
        ('Le corresponde pagar', mitad,    mitad,    total_gastado),
        ('Balance del mes',      balance_p1, balance_p2, 0),
    ]
    for i, (label, v1, v2, vt) in enumerate(filas_bal):
        ws.append([label, v1, v2, vt if label != 'Balance del mes' else ''])
        r = ws.max_row
        is_total = label == 'Balance del mes'
        row_fill = fill(COLOR_TOTAL) if is_total else fill(COLOR_WHITE if i % 2 == 0 else COLOR_ALT)
        for c in range(1, 5):
            cell = ws.cell(row=r, column=c)
            cell.font  = body_font(bold=is_total)
            cell.fill  = row_fill
            cell.border = BORDER_THIN
            if c > 1 and cell.value != '':
                cell.number_format = money_fmt()
                cell.alignment = center()
            else:
                cell.alignment = vcenter()
        # Color balance
        if is_total:
            for col, val in [(2, balance_p1), (3, balance_p2)]:
                cell = ws.cell(row=r, column=col)
                cell.font = Font(name='Arial', bold=True, size=10,
                                 color=('1E7E34' if val >= 0 else 'C0392B'))

    ws.append([])

    # Nota deuda
    if balance_p1 < 0 or balance_p2 < 0:
        deudor   = n1 if balance_p1 < 0 else n2
        acreedor = n2 if balance_p1 < 0 else n1
        monto    = abs(balance_p1 if balance_p1 < 0 else balance_p2)
        ws.merge_cells(f'A{ws.max_row + 1}:D{ws.max_row + 1}')
        nota_row = ws.max_row
        ws[f'A{nota_row}'] = f'💸  {deudor} le debe ${monto:,.2f} a {acreedor} este mes'
        ws[f'A{nota_row}'].font  = Font(name='Arial', bold=True, color='856404', size=10)
        ws[f'A{nota_row}'].fill  = fill('FFF3CD')
        ws[f'A{nota_row}'].alignment = vcenter()
        ws.row_dimensions[nota_row].height = 22

    # ── Hoja 2: Gastos del mes ─────────────────────────────────────────────────
    ws2 = wb.create_sheet('Gastos')
    ws2.sheet_view.showGridLines = False
    col_widths = [32, 18, 14, 14, 14, 12, 20]
    col_names  = ['A','B','C','D','E','F','G']
    for col, w in zip(col_names, col_widths):
        ws2.column_dimensions[col].width = w

    ws2.merge_cells('A1:G1')
    ws2['A1'] = f'Gastos aplicados — {mes_label}'
    ws2['A1'].font = Font(name='Arial', bold=True, color=COLOR_WHITE, size=13)
    ws2['A1'].fill = fill(COLOR_HEADER)
    ws2['A1'].alignment = center()
    ws2.row_dimensions[1].height = 28

    headers_g = ['Descripción', 'Categoría', 'Pagó', 'Monto total', 'Cuota mes', 'Meses', 'Mes inicio']
    ws2.append(headers_g)
    r = ws2.max_row
    for c, h in enumerate(headers_g, 1):
        cell = ws2.cell(row=r, column=c, value=h)
        cell.font = hdr_font()
        cell.fill = fill(COLOR_SUBHEADER)
        cell.alignment = center()
        cell.border = BORDER_THIN
    ws2.row_dimensions[r].height = 20

    for i, g in enumerate(gastos):
        pagador = n1 if g['pagado_por'] == 'persona1' else n2
        ws2.append([
            g['descripcion'], g['categoria'], pagador,
            g['monto_total'], g['cuota_mes'], g['meses_diferidos'],
            g['mes_inicio'].strftime('%b %Y') if hasattr(g['mes_inicio'], 'strftime') else str(g['mes_inicio'])
        ])
        r = ws2.max_row
        row_fill = fill(COLOR_WHITE if i % 2 == 0 else COLOR_ALT)
        for c in range(1, 8):
            cell = ws2.cell(row=r, column=c)
            cell.font  = body_font()
            cell.fill  = row_fill
            cell.border = BORDER_THIN
            if c in (4, 5):
                cell.number_format = money_fmt()
                cell.alignment = center()
            elif c == 6:
                cell.alignment = center()
            else:
                cell.alignment = vcenter()

    # Fila total
    if gastos:
        ws2.append(['TOTAL', '', '', f'=SUM(D3:D{ws2.max_row})', f'=SUM(E3:E{ws2.max_row})', '', ''])
        r = ws2.max_row
        for c in range(1, 8):
            cell = ws2.cell(row=r, column=c)
            cell.font  = body_font(bold=True)
            cell.fill  = fill(COLOR_TOTAL)
            cell.border = BORDER_THIN
            if c in (4, 5):
                cell.number_format = money_fmt()
                cell.alignment = center()
            else:
                cell.alignment = vcenter()

    # ── Hoja 3: Pagos Extra ────────────────────────────────────────────────────
    ws3 = wb.create_sheet('Pagos Extra')
    ws3.sheet_view.showGridLines = False
    for col, w in zip(['A','B','C','D','E'], [32, 16, 16, 14, 28]):
        ws3.column_dimensions[col].width = w

    ws3.merge_cells('A1:E1')
    ws3['A1'] = f'Pagos Extra — {mes_label}'
    ws3['A1'].font = Font(name='Arial', bold=True, color=COLOR_WHITE, size=13)
    ws3['A1'].fill = fill(COLOR_HEADER)
    ws3['A1'].alignment = center()
    ws3.row_dimensions[1].height = 28

    headers_p = ['Descripción', 'Pagó', 'Recibió', 'Monto', 'Notas']
    ws3.append(headers_p)
    r = ws3.max_row
    for c, h in enumerate(headers_p, 1):
        cell = ws3.cell(row=r, column=c, value=h)
        cell.font = hdr_font()
        cell.fill = fill(COLOR_SUBHEADER)
        cell.alignment = center()
        cell.border = BORDER_THIN
    ws3.row_dimensions[r].height = 20

    if pagos_extra:
        for i, p in enumerate(pagos_extra):
            pagador  = n1 if p['pagado_por']  == 'persona1' else n2
            receptor = n1 if p['recibido_por'] == 'persona1' else n2
            ws3.append([p['descripcion'], pagador, receptor, float(p['monto']), p['notas'] or ''])
            r = ws3.max_row
            row_fill = fill(COLOR_WHITE if i % 2 == 0 else COLOR_ALT)
            for c in range(1, 6):
                cell = ws3.cell(row=r, column=c)
                cell.font  = body_font()
                cell.fill  = row_fill
                cell.border = BORDER_THIN
                if c == 4:
                    cell.number_format = money_fmt()
                    cell.alignment = center()
                else:
                    cell.alignment = vcenter()
        # Total
        ws3.append(['TOTAL', '', '', f'=SUM(D3:D{ws3.max_row})', ''])
        r = ws3.max_row
        for c in range(1, 6):
            cell = ws3.cell(row=r, column=c)
            cell.font  = body_font(bold=True)
            cell.fill  = fill(COLOR_TOTAL)
            cell.border = BORDER_THIN
            if c == 4:
                cell.number_format = money_fmt()
                cell.alignment = center()
    else:
        ws3.merge_cells('A3:E3')
        ws3['A3'] = 'Sin pagos extra este mes'
        ws3['A3'].font = Font(name='Arial', italic=True, color='6B6860', size=10)
        ws3['A3'].alignment = center()

    return wb


# ── Rutas ──────────────────────────────────────────────────────────────────────

@descargas_bp.route('/')
@login_required
def index():
    cfg   = get_config()
    meses = meses_disponibles()
    hoy   = date.today()
    return render_template('descargas/index.html', cfg=cfg, meses=meses,
                           year_sel=hoy.year, month_sel=hoy.month)


@descargas_bp.route('/excel')
@login_required
def excel():
    hoy   = date.today()
    year  = int(request.args.get('year',  hoy.year))
    month = int(request.args.get('month', hoy.month))
    cfg   = get_config()

    gastos, pagos_extra = datos_mes(year, month)
    wb = build_excel(year, month, gastos, pagos_extra, cfg)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    mes_str = date(year, month, 1).strftime('%Y-%m')
    filename = f'gastos_{mes_str}.xlsx'

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
